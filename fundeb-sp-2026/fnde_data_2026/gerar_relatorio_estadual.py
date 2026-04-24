#!/usr/bin/env python3
"""
Gerador de Relatório PDF — Diagnóstico FUNDEB 2026 — REDE ESTADUAL
Potencial Máximo de Captação para governos estaduais (PR, SC, etc.)

Usa dados do fundeb_2026_br.db + Sinopse INEP Censo Escolar 2024.
Formato idêntico ao relatório municipal (gerar_relatorio_potencial.py).

Uso: python3 gerar_relatorio_estadual.py SC PR
"""

import sys
import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)
try:
    from weasyprint import HTML
except ImportError:
    print("pip install weasyprint"); sys.exit(1)
try:
    import pandas as pd
except ImportError:
    print("pip install pandas"); sys.exit(1)

BASE = Path(__file__).parent
XLSX_RECEITA = BASE / "01_receita_total_fundeb.xlsx"
XLSX_VAAR = BASE / "06_redes_beneficiadas_vaar.xlsx"
SINOPSE = BASE / "inep/sinopse_estatistica_censo_escolar_2024/Sinopse_Estatistica_da_Educação_Basica_2024.xlsx"
OUT_DIR = BASE / "relatorios_estadual"
OUT_DIR.mkdir(exist_ok=True)
FONTS_DIR = BASE / "fonts"

TODAY = datetime.now().strftime("%d/%m/%Y")
VAAF_MIN = 5962.79
VAAR_MEDIAN_PER_STUDENT = 710.24
PETI_FOMENTO = 1693.22

UF_NAMES = {
    "AC": "Acre", "AL": "Alagoas", "AM": "Amazonas", "AP": "Amapá",
    "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal", "ES": "Espírito Santo",
    "GO": "Goiás", "MA": "Maranhão", "MG": "Minas Gerais", "MS": "Mato Grosso do Sul",
    "MT": "Mato Grosso", "PA": "Pará", "PB": "Paraíba", "PE": "Pernambuco",
    "PI": "Piauí", "PR": "Paraná", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
    "RO": "Rondônia", "RR": "Roraima", "RS": "Rio Grande do Sul", "SC": "Santa Catarina",
    "SE": "Sergipe", "SP": "São Paulo", "TO": "Tocantins",
}

UF_IBGE = {
    "AC":12,"AL":27,"AM":13,"AP":16,"BA":29,"CE":23,"DF":53,"ES":32,
    "GO":52,"MA":21,"MG":31,"MS":50,"MT":51,"PA":15,"PB":25,"PE":26,
    "PI":22,"PR":41,"RJ":33,"RN":24,"RO":11,"RR":14,"RS":43,"SC":42,
    "SE":28,"SP":35,"TO":17,
}

SINOPSE_STATES = {
    "PR": "Paraná", "SC": "Santa Catarina", "SP": "São Paulo",
    "RS": "Rio Grande do Sul", "MG": "Minas Gerais", "RJ": "Rio de Janeiro",
    "BA": "Bahia", "CE": "Ceará", "PE": "Pernambuco", "PA": "Pará",
    "GO": "Goiás", "MA": "Maranhão", "AL": "Alagoas", "AM": "Amazonas",
    "PB": "Paraíba", "PI": "Piauí", "SE": "Sergipe", "RN": "Rio Grande do Norte",
    "MT": "Mato Grosso", "MS": "Mato Grosso do Sul", "ES": "Espírito Santo",
    "DF": "Distrito Federal", "TO": "Tocantins", "RO": "Rondônia",
    "AC": "Acre", "AP": "Amapá", "RR": "Roraima",
}

STATE_CATS = [
    "EM Integral",
    "EM Parcial",
    "EM Integrado Ed. Profissional",
    "Ens. Fund. Anos Finais",
    "Ens. Fund. Integral",
    "Ens. Fund. Anos Iniciais",
    "EJA Médio",
    "EJA Fundamental",
    "Ed. Especial EM",
    "Ed. Especial EF",
    "Ed. Especial EI",
]

CAT_FATORES = {
    "EM Integral":                   (1.52, 1.748, 2.128),
    "EM Parcial":                    (1.25, 1.4375, 1.75),
    "EM Integrado Ed. Profissional": (1.35, 1.5525, 1.89),
    "Ens. Fund. Integral":           (1.50, 1.725, 2.10),
    "Ens. Fund. Anos Finais":        (1.10, 1.265, 1.54),
    "Ens. Fund. Anos Iniciais":      (1.00, 1.15, 1.40),
    "EJA Médio":                     (1.00, 1.15, 1.40),
    "EJA Fundamental":               (1.00, 1.15, 1.40),
    "Ed. Especial EM":               (1.40, 1.61, 1.96),
    "Ed. Especial EF":               (1.40, 1.61, 1.96),
    "Ed. Especial EI":               (1.40, 1.61, 1.96),
}

AEE_FATOR = 1.40

CONVERSION_PAIRS = [
    ("EM Parcial", "EM Integral"),
    ("Ens. Fund. Anos Finais", "Ens. Fund. Integral"),
    ("Ens. Fund. Anos Iniciais", "Ens. Fund. Integral"),
]

ED_ESPECIAL_CATS = ["Ed. Especial EM", "Ed. Especial EF", "Ed. Especial EI"]


def safe_int(v):
    try:
        return int(v) if v else 0
    except (TypeError, ValueError):
        return 0


def fmt(v):
    if abs(v) >= 1_000_000:
        return f"R$ {v/1_000_000:,.1f}M".replace(",", "X").replace(".", ",").replace("X", ".")
    if abs(v) >= 1_000:
        return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_n(v):
    return f"{v:,}".replace(",", ".")


def load_receita(uf):
    df = pd.read_excel(XLSX_RECEITA, header=8, sheet_name=0)
    df.columns = ['uf', 'codigo_ibge', 'entidade', 'receita_contribuicao',
                   'compl_vaaf', 'compl_vaat', 'compl_vaar', 'compl_uniao_total', 'total_receitas']
    ibge = UF_IBGE[uf]
    row = df[(df['codigo_ibge'] == ibge) & (df['entidade'].astype(str).str.contains('GOVERNO', case=False, na=False))]
    if row.empty:
        print(f"ERRO: Governo do Estado {uf} não encontrado em receita_total")
        return None
    r = row.iloc[0]
    return {
        'receita_contribuicao': float(r['receita_contribuicao']),
        'compl_vaaf': float(r['compl_vaaf']),
        'compl_vaat': float(r['compl_vaat']),
        'compl_vaar': float(r['compl_vaar']),
        'compl_uniao_total': float(r['compl_uniao_total']),
        'total_receitas': float(r['total_receitas']),
    }


def load_vaar(uf):
    df = pd.read_excel(XLSX_VAAR, header=7, sheet_name=0)
    df.columns = ['uf', 'entidade', 'codigo_ibge', 'coef_vaar', 'compl_vaar']
    ibge = UF_IBGE[uf]
    row = df[df['codigo_ibge'] == ibge]
    if row.empty:
        return {'recebe_vaar': False, 'coef_vaar': 0, 'compl_vaar': 0}
    r = row.iloc[0]
    return {
        'recebe_vaar': True,
        'coef_vaar': float(r['coef_vaar']),
        'compl_vaar': float(r['compl_vaar']),
    }


def find_state_row(ws, state_name):
    for row in ws.iter_rows(min_row=10, values_only=True):
        uf_val = str(row[1]).strip() if row[1] else ''
        mun = str(row[2]).strip() if row[2] else ''
        if state_name in uf_val and (not mun or mun.strip() == ''):
            return row
    return None


def load_inep(uf):
    state_name = SINOPSE_STATES[uf]
    wb = openpyxl.load_workbook(str(SINOPSE), read_only=True, data_only=True)
    d = {}

    ws = wb['1.2']
    row = find_state_row(ws, state_name)
    if row:
        d['basica_total'] = safe_int(row[4])
        d['basica_est_urb'] = safe_int(row[7])
        d['basica_est_rur'] = safe_int(row[12])
        d['basica_est_total'] = safe_int(row[7]) + safe_int(row[12])

    ws = wb['1.29']
    row = find_state_row(ws, state_name)
    if row:
        d['em_total'] = safe_int(row[4])
        d['em_integral_est'] = safe_int(row[7])
        d['em_parcial_est'] = safe_int(row[12])

    ws = wb['1.26']
    row = find_state_row(ws, state_name)
    if row:
        d['em_urb_est'] = safe_int(row[7])
        d['em_rur_est'] = safe_int(row[12])

    ws = wb['1.16']
    row = find_state_row(ws, state_name)
    if row:
        d['ef_ai_est'] = safe_int(row[7]) + safe_int(row[12])
        d['ef_ai_urb_est'] = safe_int(row[7])
        d['ef_ai_rur_est'] = safe_int(row[12])

    ws = wb['1.21']
    row = find_state_row(ws, state_name)
    if row:
        d['ef_af_est'] = safe_int(row[7]) + safe_int(row[12])
        d['ef_af_urb_est'] = safe_int(row[7])
        d['ef_af_rur_est'] = safe_int(row[12])

    ws = wb['EJA 1.35']
    row = find_state_row(ws, state_name)
    if row:
        d['eja_fund_est'] = safe_int(row[7])
        d['eja_med_est'] = safe_int(row[12])

    ws = wb['Educação Profissional 1.30']
    row = find_state_row(ws, state_name)
    if row:
        d['prof_tec_integrado_est'] = safe_int(row[7])

    ws = wb['Classes Comuns 1.40']
    row = find_state_row(ws, state_name)
    if row:
        d['esp_total'] = safe_int(row[4])
        d['esp_ei'] = safe_int(row[5])
        d['esp_ef'] = safe_int(row[8])
        d['esp_em'] = safe_int(row[11])

    ws = wb['Educação Indígena 1.52']
    row = find_state_row(ws, state_name)
    if row:
        d['indigena_total'] = safe_int(row[4])

    # EF integral/parcial check sheets 1.19 and 1.24
    for sheet_key, prefix in [('1.19', 'ef_ai'), ('1.24', 'ef_af')]:
        ws = wb[sheet_key]
        hdr = None
        for i, r in enumerate(ws.iter_rows(min_row=6, max_row=10, values_only=True)):
            if r and 'Estadual' in str(r):
                hdr = r
                break
        row = find_state_row(ws, state_name)
        if row:
            d[f'{prefix}_int_est'] = safe_int(row[7])
            d[f'{prefix}_par_est'] = safe_int(row[12])

    wb.close()
    return d


def build_cats_state(inep):
    cats = {}
    total_mat_est = 0

    def add(name, urb, campo, ind_quilomb):
        nonlocal total_mat_est
        cats[name] = {"Urbano": urb, "Campo": campo, "Ind/Quilomb": ind_quilomb}
        total_mat_est += urb + campo + ind_quilomb

    em_urb = inep.get('em_urb_est', 0)
    em_rur = inep.get('em_rur_est', 0)
    em_int = inep.get('em_integral_est', 0)
    em_par = inep.get('em_parcial_est', 0)
    indigena = inep.get('indigena_total', 0)
    em_total_est = em_urb + em_rur

    em_rur_pct = em_rur / em_total_est if em_total_est else 0
    ind_pct = min(0.02, indigena / max(1, inep.get('basica_est_total', 1)))

    em_int_rur = int(em_int * em_rur_pct)
    em_int_ind = int(em_int * ind_pct)
    em_int_urb = em_int - em_int_rur - em_int_ind
    add("EM Integral", max(0, em_int_urb), max(0, em_int_rur - em_int_ind), max(0, em_int_ind))

    em_par_rur = int(em_par * em_rur_pct)
    em_par_ind = int(em_par * ind_pct)
    em_par_urb = em_par - em_par_rur - em_par_ind
    add("EM Parcial", max(0, em_par_urb), max(0, em_par_rur - em_par_ind), max(0, em_par_ind))

    prof = inep.get('prof_tec_integrado_est', 0)
    add("EM Integrado Ed. Profissional", prof, 0, 0)

    ef_af = inep.get('ef_af_est', 0)
    ef_af_urb = inep.get('ef_af_urb_est', 0)
    ef_af_rur = inep.get('ef_af_rur_est', 0)
    ef_af_int = inep.get('ef_af_int_est', 0)
    ef_af_par = inep.get('ef_af_par_est', 0)
    ef_af_ind = int(ef_af * ind_pct)
    add("Ens. Fund. Anos Finais", max(0, ef_af_urb - ef_af_ind), max(0, ef_af_rur), ef_af_ind)

    ef_int_est = inep.get('ef_ai_int_est', 0) + inep.get('ef_af_int_est', 0)
    add("Ens. Fund. Integral", ef_int_est, 0, 0)

    ef_ai = inep.get('ef_ai_est', 0)
    ef_ai_urb = inep.get('ef_ai_urb_est', 0)
    ef_ai_rur = inep.get('ef_ai_rur_est', 0)
    add("Ens. Fund. Anos Iniciais", ef_ai_urb, ef_ai_rur, 0)

    eja_med = inep.get('eja_med_est', 0)
    add("EJA Médio", eja_med, 0, 0)

    eja_fund = inep.get('eja_fund_est', 0)
    add("EJA Fundamental", eja_fund, 0, 0)

    esp_em = inep.get('esp_em', 0)
    esp_ef = inep.get('esp_ef', 0)
    esp_ei = inep.get('esp_ei', 0)
    add("Ed. Especial EM", esp_em, 0, 0)
    add("Ed. Especial EF", esp_ef, 0, 0)
    add("Ed. Especial EI", esp_ei, 0, 0)

    return cats, total_mat_est


def calc_potencial(cats, receita, vaar_info, inep):
    vaaf_ref = {}
    for cat, (fp_u, fp_c, fp_i) in CAT_FATORES.items():
        vaaf_ref[cat] = {
            "Urbano": round(VAAF_MIN * fp_u, 2),
            "Campo": round(VAAF_MIN * fp_c, 2),
            "Ind/Quilomb": round(VAAF_MIN * fp_i, 2),
            "fator": fp_u,
        }

    total_mat_est = sum(
        sum(locs.values()) for locs in cats.values()
    )

    receita_fundeb = receita['total_receitas']
    cats_ativas = sum(1 for c in cats.values() if sum(c.values()) > 0)
    cats_faltantes = len(STATE_CATS) - cats_ativas

    # T1: categorias com 0 matrículas
    t1_items = []
    t1_total = 0
    for cat in STATE_CATS:
        if cat not in cats:
            continue
        mat = sum(cats[cat].values())
        if mat == 0:
            vaaf = vaaf_ref[cat]["Urbano"]
            ganho_10 = round(vaaf * 10)
            ganho_50 = round(vaaf * 50)
            t1_items.append({"cat": cat, "vaaf": vaaf, "g10": ganho_10, "g50": ganho_50})
            t1_total += ganho_10

    # T2: conversão parcial → integral
    t2_items = []
    t2_total = 0
    for parcial, integral in CONVERSION_PAIRS:
        if parcial not in cats or integral not in cats:
            continue
        alunos_parcial = sum(cats[parcial].values())
        if alunos_parcial == 0:
            continue
        fp_parcial = CAT_FATORES[parcial][0]
        fp_integral = CAT_FATORES[integral][0]
        dif = round((fp_integral - fp_parcial) * VAAF_MIN, 2)
        ganho = round(alunos_parcial * dif)
        t2_items.append({
            "de": parcial, "para": integral,
            "alunos": alunos_parcial, "dif": dif, "ganho": ganho
        })
        t2_total += ganho

    # T3: AEE dupla matrícula
    t3_items = []
    t3_total = 0
    vaaf_aee = round(VAAF_MIN * AEE_FATOR, 2)
    for cat in ED_ESPECIAL_CATS:
        if cat not in cats:
            continue
        alunos = sum(cats[cat].values())
        if alunos == 0:
            continue
        ganho = round(alunos * vaaf_aee)
        t3_items.append({"cat": cat, "alunos": alunos, "vaaf_aee": vaaf_aee, "ganho": ganho})
        t3_total += ganho

    # T4: reclassificação de localidade
    mat_urbanas = sum(c.get("Urbano", 0) for c in cats.values())
    mat_campo = sum(c.get("Campo", 0) for c in cats.values())
    mat_ind = sum(c.get("Ind/Quilomb", 0) for c in cats.values())
    tem_campo = mat_campo > 0
    vaaf_medio = receita_fundeb / max(1, total_mat_est)
    t4_campo = round(mat_urbanas * 0.10 * vaaf_medio * 0.15)
    t4_ind = round(mat_urbanas * 0.05 * vaaf_medio * 0.40)
    t4_total = t4_campo + t4_ind

    # T5: VAAR
    recebe_vaar = vaar_info.get('recebe_vaar', False)
    vaar_atual = vaar_info.get('compl_vaar', 0) if recebe_vaar else 0
    vaar_potencial = round(total_mat_est * VAAR_MEDIAN_PER_STUDENT)
    t5_total = max(0, vaar_potencial - vaar_atual)

    # T6: EC 135
    t6_pct4 = round(receita_fundeb * 0.04)
    em_int = inep.get('em_integral_est', 0)
    em_total_est = inep.get('em_urb_est', 0) + inep.get('em_rur_est', 0)
    pct_integral = em_int / max(1, em_total_est) * 100
    novas_vagas = round(t6_pct4 / (VAAF_MIN * 1.52)) if t6_pct4 > 0 else 0
    t6_total = t6_pct4

    potencial_total = t1_total + t2_total + t3_total + t4_total + t5_total + t6_total
    pct_ganho = potencial_total / max(1, receita_fundeb) * 100

    return {
        "receita_fundeb": receita_fundeb,
        "potencial_total": potencial_total,
        "pct_ganho": pct_ganho,
        "total_mat_est": total_mat_est,
        "cats_ativas": cats_ativas,
        "cats_total": len(STATE_CATS),
        "cats_faltantes": cats_faltantes,
        "recebe_vaar": recebe_vaar,
        "t1": {"items": t1_items, "total": t1_total},
        "t2": {"items": t2_items, "total": t2_total},
        "t3": {"items": t3_items, "total": t3_total},
        "t4": {
            "mat_urbanas": mat_urbanas, "tem_campo": tem_campo,
            "campo": t4_campo, "ind": t4_ind, "total": t4_total
        },
        "t5": {
            "recebe": recebe_vaar, "atual": vaar_atual,
            "potencial": vaar_potencial, "total": t5_total
        },
        "t6": {
            "pct4": t6_pct4, "pct_integral": pct_integral,
            "novas_vagas": novas_vagas, "total": t6_total
        },
    }


def build_html(uf, receita, vaar_info, inep, cats, pot):
    nome_estado = UF_NAMES[uf]
    nome_upper = f"REDE ESTADUAL — {nome_estado.upper()}"

    tiers = [
        ("T1 VAAF", pot["t1"]["total"]),
        ("T2 Integral", pot["t2"]["total"]),
        ("T3 AEE", pot["t3"]["total"]),
        ("T4 Localidade", pot["t4"]["total"]),
        ("T5 VAAR", pot["t5"]["total"]),
        ("T6 EC 135", pot["t6"]["total"]),
    ]
    max_tier = max(t[1] for t in tiers) if tiers else 1

    tier_bars = ""
    for label, val in tiers:
        pct = val / max(1, max_tier) * 100
        tier_bars += f"""
        <div class="tier-row">
            <span class="tier-label">{label}</span>
            <div class="tier-bar-bg"><div class="tier-bar" style="width:{pct}%"></div></div>
            <span class="tier-value">{fmt(val)}</span>
        </div>"""

    # T1 table
    t1_rows = ""
    for item in pot["t1"]["items"]:
        t1_rows += f"""
        <tr>
            <td>{item['cat']}</td>
            <td class="right">{fmt(item['vaaf'])}</td>
            <td class="right">{fmt(item['g10'])}</td>
            <td class="right">{fmt(item['g50'])}</td>
        </tr>"""

    # T2 table
    t2_rows = ""
    for item in pot["t2"]["items"]:
        t2_rows += f"""
        <tr>
            <td>{item['de']}</td>
            <td class="accent">{item['para']}</td>
            <td class="right">{fmt_n(item['alunos'])}</td>
            <td class="right accent">+{fmt(item['dif'])}</td>
            <td class="right bold accent">{fmt(item['ganho'])}</td>
        </tr>"""

    # T3 table
    t3_rows = ""
    for item in pot["t3"]["items"]:
        t3_rows += f"""
        <tr>
            <td>{item['cat']}</td>
            <td class="right">{fmt_n(item['alunos'])}</td>
            <td class="right">{fmt(item['vaaf_aee'])}</td>
            <td class="right bold accent">{fmt(item['ganho'])}</td>
        </tr>"""

    recebe_vaar = pot["t5"]["recebe"]
    vaar_atual = pot["t5"]["atual"]
    vaar_status = "Recebe" if recebe_vaar else "Não recebe"
    vaar_class = "" if recebe_vaar else "red"

    # Recommendations
    recs = []
    sorted_tiers = sorted(tiers, key=lambda x: -x[1])
    for label, val in sorted_tiers:
        if val <= 0:
            continue
        tier_key = label.split()[0]
        if tier_key == "T2":
            recs.append(("ALTO", "T2", "Converter Parcial para Integral no Ensino Médio",
                f"{fmt_n(pot['t2']['items'][0]['alunos'] if pot['t2']['items'] else 0)} alunos do EM em jornada parcial. A conversão representa um ganho de {fmt(val)}/ano. A EC 135 exige 4% do FUNDEB para novas vagas integrais."))
        elif tier_key == "T3":
            total_esp = sum(i['alunos'] for i in pot['t3']['items'])
            recs.append(("ALTO", "T3", "Maximizar o AEE (Dupla Matrícula)",
                f"{fmt_n(total_esp)} alunos em Educação Especial. Se todos tiverem o AEE registrado, o ganho é de {fmt(val)}/ano."))
        elif tier_key == "T5":
            if not pot['t5']['recebe']:
                recs.append(("ALTO", "T5", "Cumprir as 5 Condicionalidades do VAAR",
                    f"Não recebe VAAR. Potencial: {fmt(val)}/ano. Requer: gestores selecionados por mérito, SAEB ≥ 80%, BNCC Computação e ICMS Educacional."))
            else:
                recs.append(("MEDIO", "T5", "Maximizar o VAAR",
                    f"Já recebe VAAR ({fmt(pot['t5']['atual'])}). Potencial adicional de {fmt(val)}/ano com a otimização dos indicadores."))
        elif tier_key == "T6":
            recs.append(("ALTO", "T6", "Implementar a BNCC Computação e Expandir o Integral",
                f"Obrigatório a partir de 2026. Condicionalidade V do VAAR. 4% do FUNDEB = {fmt(val)} para expansão do tempo integral."))
        elif tier_key == "T1":
            recs.append(("MEDIO", "T1", f"Abrir {pot['t1']['items'].__len__()} Categorias Faltantes",
                f"Categorias sem matrícula representam receita não captada. Potencial: {fmt(val)}/ano."))
        elif tier_key == "T4":
            recs.append(("MEDIO", "T4", "Reclassificar Escolas (Campo / Indígena / Quilombola)",
                f"Escolas rurais ou em terras indígenas com classificação incorreta no Educacenso. Potencial: {fmt(val)}/ano."))

    recs_html = ""
    for prio, tier, title, desc in recs:
        color = "#e53e3e" if prio == "ALTO" else "#d69e2e"
        recs_html += f"""
        <div class="rec-item">
            <span class="rec-prio" style="border-color:{color};color:{color}">{prio}</span>
            <span class="rec-tier">{tier}</span>
            <strong>{title}</strong>
            <p class="rec-desc">{desc}</p>
        </div>"""

    # Comparison with other states
    comp_data = load_state_comparison(uf)
    comp_rows = ""
    for row in comp_data:
        is_self = row['uf'] == uf
        bold = "font-weight:700;" if is_self else ""
        suffix = " (você)" if is_self else ""
        comp_rows += f"""
        <tr style="{bold}">
            <td>REDE ESTADUAL {row['uf']}{suffix}</td>
            <td class="right">{fmt_n(row['mat'])}</td>
            <td class="right">{fmt(row['receita'])}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<style>
@page {{ size: A4; margin: 1.5cm 2cm; }}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: 'Segoe UI', system-ui, sans-serif; color: #2d3748; font-size: 10pt; line-height: 1.5; }}

.cover {{ background: linear-gradient(135deg, #1a365d, #2b6cb0); color: white; min-height: 100vh; padding: 3rem 2.5rem; page-break-after: always; display: flex; flex-direction: column; justify-content: space-between; }}
.cover .brand {{ color: #4fd1c5; font-size: 14pt; font-weight: 700; margin-bottom: 2rem; }}
.cover .subtitle {{ font-size: 10pt; color: rgba(255,255,255,0.8); margin-bottom: 3rem; }}
.cover .entity {{ font-size: 28pt; font-weight: 800; margin-bottom: 0.5rem; }}
.cover .entity-sub {{ font-size: 10pt; color: rgba(255,255,255,0.7); margin-bottom: 4rem; }}
.cover .big-number {{ font-size: 48pt; font-weight: 800; color: #48bb78; margin-bottom: 0.5rem; }}
.cover .big-label {{ font-size: 11pt; color: rgba(255,255,255,0.8); margin-bottom: 2rem; }}
.cover .pct-badge {{ display: inline-block; border: 2px solid #4fd1c5; color: #4fd1c5; padding: 0.5rem 2rem; border-radius: 8px; font-size: 14pt; font-weight: 700; }}
.cover .footer {{ font-size: 8pt; color: rgba(255,255,255,0.5); margin-top: 2rem; }}
.cover .footer .brand-small {{ color: #4fd1c5; font-weight: 700; }}

.page {{ page-break-before: always; padding-top: 0.5rem; }}
h2 {{ font-size: 16pt; color: #1a365d; margin-bottom: 1rem; }}
h3 {{ font-size: 12pt; color: #2b6cb0; margin: 1.5rem 0 0.5rem; }}

.metrics {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 0.75rem; margin-bottom: 1.5rem; }}
.metric {{ border: 1px solid #e2e8f0; border-radius: 6px; padding: 0.75rem; text-align: center; }}
.metric .label {{ font-size: 7pt; text-transform: uppercase; color: #718096; letter-spacing: 0.5px; }}
.metric .value {{ font-size: 16pt; font-weight: 700; color: #1a365d; }}
.metric .value.accent {{ color: #2b6cb0; }}
.metric .value.red {{ color: #e53e3e; }}
.metric .value.green {{ color: #38a169; }}

.tier-row {{ display: flex; align-items: center; margin: 0.4rem 0; }}
.tier-label {{ width: 100px; font-size: 9pt; font-weight: 600; color: #4a5568; }}
.tier-bar-bg {{ flex: 1; height: 20px; background: #edf2f7; border-radius: 4px; overflow: hidden; margin: 0 0.5rem; }}
.tier-bar {{ height: 100%; background: linear-gradient(90deg, #4fd1c5, #2b6cb0); border-radius: 4px; }}
.tier-value {{ width: 100px; text-align: right; font-size: 9pt; font-weight: 700; color: #1a365d; }}
.tier-total {{ font-size: 14pt; font-weight: 800; color: #2b6cb0; margin-top: 0.75rem; }}

.alert {{ border: 1.5px solid #e53e3e; border-radius: 8px; padding: 0.75rem 1rem; margin-top: 1rem; font-size: 9pt; color: #4a5568; }}
.alert strong {{ color: #e53e3e; }}

table {{ width: 100%; border-collapse: collapse; margin: 0.5rem 0; font-size: 9pt; }}
th {{ text-transform: uppercase; font-size: 7.5pt; letter-spacing: 0.5px; color: #718096; border-bottom: 2px solid #e2e8f0; padding: 0.4rem 0.5rem; text-align: left; }}
td {{ padding: 0.4rem 0.5rem; border-bottom: 1px solid #edf2f7; }}
.right {{ text-align: right; }}
.bold {{ font-weight: 700; }}
.accent {{ color: #2b6cb0; }}
.red {{ color: #e53e3e; }}
.green {{ color: #38a169; }}

.t5-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; margin: 0.5rem 0; }}
.t5-item {{ display: flex; justify-content: space-between; font-size: 9pt; padding: 0.2rem 0; }}
.t5-label {{ color: #718096; }}
.t5-val {{ font-weight: 700; }}

.rec-item {{ border-left: 3px solid #e2e8f0; padding: 0.5rem 0.75rem; margin: 0.5rem 0; }}
.rec-prio {{ display: inline-block; border: 1.5px solid; border-radius: 3px; padding: 0.1rem 0.4rem; font-size: 7pt; font-weight: 700; margin-right: 0.3rem; }}
.rec-tier {{ display: inline-block; background: #edf2f7; border-radius: 3px; padding: 0.1rem 0.4rem; font-size: 7pt; font-weight: 600; color: #4a5568; margin-right: 0.3rem; }}
.rec-desc {{ font-size: 8.5pt; color: #718096; margin-top: 0.2rem; }}

.phase {{ border-left: 3px solid #4fd1c5; padding: 0.5rem 0.75rem; margin: 0.75rem 0; }}
.phase .phase-title {{ display: inline-block; background: #e6fffa; color: #234e52; padding: 0.15rem 0.5rem; border-radius: 4px; font-size: 9pt; font-weight: 600; margin-bottom: 0.3rem; }}
.phase li {{ font-size: 9pt; margin: 0.2rem 0; margin-left: 1rem; }}

.cta {{ background: linear-gradient(135deg, #1a365d, #2b6cb0); color: white; border-radius: 12px; padding: 1.5rem 2rem; text-align: center; margin-top: 2rem; }}
.cta h3 {{ color: #4fd1c5; font-size: 12pt; margin-bottom: 0.3rem; }}
.cta p {{ font-size: 9pt; color: rgba(255,255,255,0.8); margin-bottom: 1rem; }}
.cta .services {{ display: flex; justify-content: center; gap: 1rem; flex-wrap: wrap; margin-bottom: 0.75rem; }}
.cta .svc {{ border: 1px solid rgba(255,255,255,0.3); border-radius: 6px; padding: 0.3rem 0.75rem; font-size: 8pt; }}
.cta .email {{ color: #4fd1c5; font-weight: 700; font-size: 10pt; }}

.page-footer {{ text-align: center; font-size: 7.5pt; color: #a0aec0; margin-top: 2rem; padding-top: 0.5rem; border-top: 1px solid #e2e8f0; }}

.ctx {{ border-radius: 8px; padding: 0.75rem 1rem; margin: 0.5rem 0 0.75rem; font-size: 8.5pt; line-height: 1.6; }}
.ctx-sit {{ background: #f7fafc; border-left: 4px solid #a0aec0; }}
.ctx-pot {{ background: #ebf8ff; border-left: 4px solid #2b6cb0; }}
.ctx-acao {{ border-left: 4px solid #38a169; padding: 0.5rem 1rem; }}
.ctx-curto {{ background: #f0fff4; }}
.ctx-medio {{ background: #fffff0; }}
.ctx-longo {{ background: #fff5f5; }}
.ctx strong {{ font-size: 8pt; text-transform: uppercase; letter-spacing: 0.3px; }}
.ctx .tag-prazo {{ display: inline-block; border-radius: 3px; padding: 0.1rem 0.5rem; font-size: 7pt; font-weight: 700; color: white; margin-left: 0.3rem; }}
.tag-curto {{ background: #38a169; }}
.tag-medio {{ background: #d69e2e; }}
.tag-longo {{ background: #e53e3e; }}
</style>
</head>
<body>

<!-- PAGE 1: COVER -->
<div class="cover">
    <div>
        <div class="brand">INSTITUTO i10</div>
        <div class="subtitle">Diagnóstico FUNDEB 2026 · Potencial Máximo de Captação · REDE ESTADUAL</div>
        <div class="entity">{nome_upper}</div>
        <div class="entity-sub">{fmt_n(pot['total_mat_est'])} matrículas na rede estadual</div>
    </div>
    <div>
        <div class="big-number">{fmt(pot['potencial_total'])}</div>
        <div class="big-label">em recursos FUNDEB que a Rede Estadual de {nome_estado} não está captando</div>
        <div><span class="pct-badge">+{pot['pct_ganho']:.1f}% sobre a receita atual</span></div>
    </div>
    <div class="footer">
        <span class="brand-small">Instituto i10</span> — Orquestrando o Futuro da Educação Pública<br>
        Relatório gerado em {TODAY} · Dados: MEC/FNDE FUNDEB 2026 (parâmetros definitivos) + INEP Censo 2024
    </div>
</div>

<!-- PAGE 2: RESUMO EXECUTIVO -->
<div class="page">
    <h2>Resumo Executivo</h2>
    <div class="metrics">
        <div class="metric"><div class="label">Receita FUNDEB Atual</div><div class="value">{fmt(pot['receita_fundeb'])}</div></div>
        <div class="metric"><div class="label">Potencial Não Captado</div><div class="value accent">{fmt(pot['potencial_total'])}</div></div>
        <div class="metric"><div class="label">Ganho Percentual</div><div class="value green">+{pot['pct_ganho']:.1f}%</div></div>
        <div class="metric"><div class="label">Complementação VAAR</div><div class="value {vaar_class}">{vaar_status}</div></div>
    </div>
    <div class="metrics">
        <div class="metric"><div class="label">Matrículas Rede Estadual</div><div class="value">{fmt_n(pot['total_mat_est'])}</div></div>
        <div class="metric"><div class="label">Categorias Ativas</div><div class="value">{pot['cats_ativas']} / {pot['cats_total']}</div></div>
        <div class="metric"><div class="label">Categorias Faltantes</div><div class="value red">{pot['cats_faltantes']}</div></div>
        <div class="metric"><div class="label">% Ensino Médio Integral</div><div class="value">{pot['t6']['pct_integral']:.1f}%</div></div>
    </div>

    <h3>Decomposição do Potencial por Alavanca</h3>
    {tier_bars}
    <div class="tier-total">TOTAL &nbsp; {fmt(pot['potencial_total'])}</div>

    <div class="alert">
        <strong>ALERTA CRÍTICO:</strong> A Rede Estadual de {nome_estado} {"não recebe" if not pot['recebe_vaar'] else "recebe"} complementação VAAR
        (R$ 7,5 bilhões disponíveis nacionalmente).
        {"Potencial estimado: " + fmt(pot['t5']['total']) + "/ano se cumprir as 5 condicionalidades MEC." if not pot['recebe_vaar'] else "Valor atual: " + fmt(pot['t5']['atual']) + "/ano. Potencial adicional: " + fmt(pot['t5']['total']) + "."}
    </div>

    <div class="page-footer">Instituto i10 — Orquestrando o Futuro da Educação Pública &nbsp;&nbsp; 2 / 7</div>
</div>

<!-- PAGE 3: DETALHAMENTO T1-T3 -->
<div class="page">
    <h2>Detalhamento por Alavanca</h2>

    <h3>T1 — Expansão VAAF: {len(pot['t1']['items'])} Categorias Não Captadas</h3>
    <div class="ctx ctx-sit">
        <strong>Situação Atual:</strong> {f"A rede estadual de {nome_estado} opera em {pot['cats_ativas']} das {pot['cats_total']} categorias possíveis. Nenhuma categoria está completamente zerada, o que indica cobertura mínima em todas as etapas (Ensino Médio, Ensino Fundamental, EJA, Educação Especial e Educação Profissional)." if pot['cats_faltantes'] == 0 else f"A rede estadual de {nome_estado} opera em apenas {pot['cats_ativas']} das {pot['cats_total']} categorias possíveis. Há {pot['cats_faltantes']} categorias com ZERO matrículas — cada uma representa receita do FUNDEB que o estado contribui, mas não recebe de volta."}
    </div>
    <div class="ctx ctx-pot">
        <strong>Potencial Não Captado: {fmt(pot['t1']['total'])}</strong><br>
        {"Mesmo com todas as categorias ativas, há oportunidade de expandir aquelas com pouquíssimas matrículas. Categorias com menos de 100 alunos podem ser ampliadas com baixo custo operacional." if pot['cats_faltantes'] == 0 else f"Cada categoria aberta com apenas 10 alunos já gera receita imediata. Com 50 alunos por categoria, o ganho total estimado chega a {fmt(pot['t1']['total'])}/ano."}
    </div>
    <div class="ctx ctx-acao ctx-curto">
        <strong>Como Captar</strong> <span class="tag-prazo tag-curto">CURTO PRAZO</span><br>
        Identificar categorias sub-representadas no Educacenso. Abrir turmas ou formalizar parcerias conveniadas antes do Censo Escolar (28/mai/2026). São necessárias apenas 10 matrículas por categoria para iniciar a captação.
    </div>
    <table>
        <tr><th>Categoria Faltante</th><th class="right">VAAF/Aluno</th><th class="right">Ganho +10</th><th class="right">Ganho +50</th></tr>
        {t1_rows}
    </table>

    <h3>T2 — Conversão de Parcial para Integral</h3>
    <div class="ctx ctx-sit">
        <strong>Situação Atual:</strong> A rede estadual de {nome_estado} mantém apenas {fmt_n(inep.get('em_integral_est',0))} alunos em tempo integral no Ensino Médio — apenas {pot['t6']['pct_integral']:.1f}% do total. Isso significa que {100 - pot['t6']['pct_integral']:.1f}% dos estudantes do EM frequentam jornada parcial (meio período). No Ensino Fundamental estadual, o cenário é semelhante. Para referência, a média da região Nordeste já ultrapassa 20% de matrículas integrais — {nome_estado} está muito abaixo desse patamar.
    </div>
    <div class="ctx ctx-pot">
        <strong>Potencial Não Captado: {fmt(pot['t2']['total'])}</strong><br>
        Cada aluno convertido de EM Parcial (fator 1,25) para EM Integral (fator 1,52) gera um acréscimo de R$ 1.610 por aluno/ano. No Ensino Fundamental, a conversão de parcial para integral chega a gerar até R$ 2.981 por aluno/ano. Esta é a maior oportunidade de captação da rede estadual.
    </div>
    <div class="ctx ctx-acao ctx-medio">
        <strong>Como Captar</strong> <span class="tag-prazo tag-medio">MÉDIO PRAZO</span><br>
        Aderir ao PETI (Programa Escola em Tempo Integral) do MEC, que oferece R$ 1.693/aluno como fomento federal. Ampliar a jornada escolar para 7 horas diárias ou mais, com atividades complementares. Priorizar escolas cuja infraestrutura já seja adequada. O simples registro no Educacenso como "integral" ativa automaticamente o fator mais alto no FUNDEB do ano seguinte.
    </div>
    <table>
        <tr><th>De (Parcial)</th><th>Para (Integral)</th><th class="right">Alunos</th><th class="right">Dif./Aluno</th><th class="right">Ganho Total</th></tr>
        {t2_rows}
    </table>
    <p class="bold" style="margin-top:0.3rem">Subtotal T2 &nbsp; <span class="accent">{fmt(pot['t2']['total'])}</span></p>

    <h3>T3 — AEE e Dupla Matrícula (Educação Especial)</h3>
    <div class="ctx ctx-sit">
        <strong>Situação Atual:</strong> A rede estadual registra {fmt_n(sum(i['alunos'] for i in pot['t3']['items']))} alunos em Educação Especial (classes comuns), distribuídos entre Ensino Médio, Fundamental e Educação Infantil. Entretanto, para que esses alunos gerem a chamada "dupla matrícula" no FUNDEB, é necessário que cada um possua: (1) laudo médico registrado, (2) atendimento AEE efetivo em sala de recursos multifuncionais e (3) registro correto no Educacenso como aluno com AEE. Na prática, muitos estados incluem o aluno na classe comum, mas NÃO registram o AEE — perdendo assim o fator adicional de 1,40.
    </div>
    <div class="ctx ctx-pot">
        <strong>Potencial Não Captado: {fmt(pot['t3']['total'])}</strong><br>
        Cada aluno com deficiência que tenha o AEE registrado gera uma segunda contagem de {fmt(round(VAAF_MIN * 1.40))}/ano, além da matrícula regular. Não é necessário matricular novos alunos — basta regularizar o registro daqueles que já frequentam a rede.
    </div>
    <div class="ctx ctx-acao ctx-curto">
        <strong>Como Captar</strong> <span class="tag-prazo tag-curto">CURTO PRAZO</span><br>
        Realizar busca ativa de alunos com deficiência que estejam sem laudo ou sem AEE registrado. Equipar salas de recursos multifuncionais nas escolas estaduais. Garantir que o Educacenso registre tanto a matrícula regular quanto o atendimento AEE de cada aluno elegível. Prazo: antes de 28/mai/2026 (data de referência do Censo Escolar).
    </div>
    <table>
        <tr><th>Categoria</th><th class="right">Alunos Ed. Esp.</th><th class="right">VAAF AEE/Aluno</th><th class="right">Ganho Adicional</th></tr>
        {t3_rows}
    </table>
    <p class="bold" style="margin-top:0.3rem">Subtotal T3 &nbsp; <span class="accent">{fmt(pot['t3']['total'])}</span></p>

    <div class="page-footer">Instituto i10 — Orquestrando o Futuro da Educação Pública &nbsp;&nbsp; 3 / 7</div>
</div>

<!-- PAGE 4: T4 -->
<div class="page">
    <h3>T4 — Reclassificação de Localidade</h3>
    <div class="ctx ctx-sit">
        <strong>Situação Atual:</strong> A rede estadual de {nome_estado} registra {fmt_n(pot['t4']['mat_urbanas'])} matrículas classificadas como "urbanas" e {fmt_n(sum(c.get('Campo',0) for c in cats.values()))} como "campo" (rural). No FUNDEB, escolas em área rural recebem automaticamente um acréscimo de 15% sobre o fator de ponderação, enquanto escolas em terras indígenas ou quilombolas recebem 40% a mais. Diversas escolas estaduais localizadas em perímetros rurais, distritos ou nas proximidades de comunidades tradicionais encontram-se classificadas incorretamente como urbanas no Educacenso — deixando de captar esses multiplicadores.
    </div>
    <div class="ctx ctx-pot">
        <strong>Potencial Não Captado: {fmt(pot['t4']['total'])}</strong><br>
        Se apenas 10% das matrículas hoje classificadas como urbanas forem corretamente reclassificadas como "campo", o ganho é de {fmt(pot['t4']['campo'])}/ano. Se 5% forem reclassificadas como indígena ou quilombola, o ganho adicional alcança {fmt(pot['t4']['ind'])}/ano.
    </div>
    <div class="ctx ctx-acao ctx-curto">
        <strong>Como Captar</strong> <span class="tag-prazo tag-curto">CURTO PRAZO</span><br>
        Auditar a classificação de localidade de cada escola estadual no Educacenso. Cruzar os dados com o IBGE (setores censitários rurais) e com a FUNAI e a Fundação Palmares (terras indígenas e territórios quilombolas). Corrigir as classificações incorretas antes do Censo Escolar 2026. Não requer investimento financeiro — apenas correção cadastral.
    </div>
    <table>
        <tr><td>Matrículas urbanas</td><td class="right bold">{fmt_n(pot['t4']['mat_urbanas'])}</td><td>Tem matrículas campo</td><td class="right bold">{"Sim" if pot['t4']['tem_campo'] else "Não"}</td></tr>
        <tr><td>Ganho se 10% reclassificadas como Campo (+15%)</td><td colspan="3" class="right bold accent">{fmt(pot['t4']['campo'])}</td></tr>
        <tr><td>Ganho se 5% reclassificadas como Ind/Quilomb (+40%)</td><td colspan="3" class="right bold accent">{fmt(pot['t4']['ind'])}</td></tr>
    </table>

    <h3>T5 — Complementação Federal (VAAR + VAAT)</h3>
    <div class="ctx ctx-sit">
        <strong>Situação Atual:</strong> {"A rede estadual de " + nome_estado + " NÃO recebe a complementação VAAR atualmente. Isso indica que o estado não cumpre — ou não registrou no SIMEC — as 5 condicionalidades exigidas pelo MEC: (I) seleção de diretores por critérios de mérito, (II) participação mínima de 80% dos alunos no SAEB, (III) redução de desigualdades de aprendizagem, (IV) ICMS Educacional — exclusiva de redes estaduais, e (V) implementação da BNCC, incluindo o módulo de Computação. São R$ 7,5 bilhões disponíveis nacionalmente que " + nome_estado + " não está acessando." if not recebe_vaar else "A rede estadual de " + nome_estado + " já recebe a complementação VAAR no valor de " + fmt(vaar_atual) + "/ano. Contudo, há potencial para ampliar esse valor por meio da otimização dos indicadores de atendimento e aprendizagem que determinam o coeficiente de distribuição."}
    </div>
    <div class="ctx ctx-pot">
        <strong>Potencial Não Captado: {fmt(pot['t5']['total'])}</strong><br>
        {"Com base na mediana nacional de R$ 710 por aluno (VAAR) e nas " + fmt_n(pot['total_mat_est']) + " matrículas da rede estadual, o potencial estimado é de " + fmt(pot['t5']['potencial']) + "/ano. Esse recurso vem integralmente da União — não sai do caixa do estado." if not recebe_vaar else "Potencial adicional de " + fmt(pot['t5']['total']) + "/ano com a melhoria dos indicadores de resultado."}
    </div>
    <div class="ctx ctx-acao {"ctx-medio" if not recebe_vaar else "ctx-curto"}">
        <strong>Como Captar</strong> <span class="tag-prazo {"tag-medio" if not recebe_vaar else "tag-curto"}">{"MÉDIO PRAZO" if not recebe_vaar else "CURTO PRAZO"}</span><br>
        {"Registrar as 5 condicionalidades no SIMEC (módulo VAAR/FUNDEB). Atenção especial à Condicionalidade IV (ICMS Educacional), que é exclusiva das redes estaduais. Implementar a BNCC Computação (Condicionalidade V) com os 3 eixos obrigatórios: Pensamento Computacional, Mundo Digital e Cultura Digital. Garantir participação mínima de 80% dos alunos no SAEB. O prazo habitual para registro é agosto a setembro do ano corrente." if not recebe_vaar else "Melhorar os indicadores do SAEB, ampliar a participação dos alunos nas avaliações e garantir que a Condicionalidade IV (ICMS Educacional) e a Condicionalidade V (BNCC Computação) estejam plenamente atendidas para maximizar o coeficiente VAAR."}
    </div>
    <div class="t5-grid">
        <div>
            <div class="t5-item"><span class="t5-label">VAAR atual</span><span class="t5-val {vaar_class}">{fmt(vaar_atual) if recebe_vaar else "Não recebe"}</span></div>
            <div class="t5-item"><span class="t5-label">VAAR potencial</span><span class="t5-val accent">{fmt(pot['t5']['potencial'])}</span></div>
        </div>
        <div>
            <div class="t5-item"><span class="t5-label">VAAT atual</span><span class="t5-val">Não recebe</span></div>
            <div class="t5-item"><span class="t5-label">VAAT potencial</span><span class="t5-val accent">{fmt(round(pot['total_mat_est'] * 500))}</span></div>
        </div>
    </div>

    <div class="page-footer">Instituto i10 — Orquestrando o Futuro da Educação Pública &nbsp;&nbsp; 4 / 7</div>
</div>

<!-- PAGE 5: T6 -->
<div class="page">
    <h3>T6 — EC 135/2024 + BNCC Computação</h3>
    <div class="ctx ctx-sit">
        <strong>Situação Atual:</strong> A Emenda Constitucional 135/2024 determina que 4% da receita do FUNDEB de cada ente federado deve ser destinada obrigatoriamente à criação de novas vagas em tempo integral. Para a rede estadual de {nome_estado}, isso equivale a {fmt(pot['t6']['pct4'])} por ano. Atualmente, apenas {pot['t6']['pct_integral']:.1f}% dos alunos do Ensino Médio frequentam o regime integral ({fmt_n(inep.get('em_integral_est',0))} alunos). A meta do PNE é alcançar 50%. Além disso, a BNCC Computação tornou-se obrigatória a partir de 2026 e é uma das condicionalidades (V) do VAAR.
    </div>
    <div class="ctx ctx-pot">
        <strong>Potencial Não Captado: {fmt(pot['t6']['total'])}</strong><br>
        Os 4% obrigatórios ({fmt(pot['t6']['pct4'])}) permitem criar aproximadamente {fmt_n(pot['t6']['novas_vagas'])} novas vagas em tempo integral. Cada vaga criada passa a contar com o fator integral (1,52 para o Ensino Médio) no FUNDEB do ano seguinte, além de R$ 1.693/aluno de fomento federal via PETI. Caso o estado não expanda o tempo integral, esses recursos ficam sem destinação legal e podem gerar irregularidades na prestação de contas.
    </div>
    <div class="ctx ctx-acao ctx-medio">
        <strong>Como Captar</strong> <span class="tag-prazo tag-medio">MÉDIO PRAZO</span><br>
        Aderir ao PETI (Programa Escola em Tempo Integral) do MEC. Selecionar escolas com infraestrutura adequada para jornada ampliada (refeitório, espaços para atividades). Implementar a BNCC Computação em todas as escolas estaduais — é obrigatória e está vinculada ao recebimento do VAAR. Registrar a expansão no Educacenso e comprovar a execução no SIMEC.
    </div>
    <div class="t5-grid">
        <div>
            <div class="t5-item"><span class="t5-label">4% FUNDEB (obrigatório p/ integral)</span><span class="t5-val accent">{fmt(pot['t6']['pct4'])}</span></div>
            <div class="t5-item"><span class="t5-label">Novas vagas possíveis com 4%</span><span class="t5-val accent">~{fmt_n(pot['t6']['novas_vagas'])} alunos</span></div>
        </div>
        <div>
            <div class="t5-item"><span class="t5-label">Tempo integral atual</span><span class="t5-val">{fmt_n(inep.get('em_integral_est',0))} ({pot['t6']['pct_integral']:.1f}%)</span></div>
            <div class="t5-item"><span class="t5-label">PETI fomento federal/aluno</span><span class="t5-val">{fmt(PETI_FOMENTO)}</span></div>
        </div>
    </div>

    <div class="page-footer">Instituto i10 — Orquestrando o Futuro da Educação Pública &nbsp;&nbsp; 5 / 7</div>
</div>

<!-- PAGE 6: COMPARAÇÃO + RECOMENDAÇÕES -->
<div class="page">
    <h2>Comparação com Estados da Região Sul</h2>
    <p style="font-size:9pt;color:#718096;margin-bottom:0.5rem">Redes estaduais da mesma região:</p>
    <table>
        <tr><th>Rede Estadual</th><th class="right">Matrículas</th><th class="right">Receita FUNDEB</th></tr>
        {comp_rows}
    </table>

    <h2 style="margin-top:1.5rem">Recomendações Priorizadas</h2>
    {recs_html}

    <h2 style="margin-top:1.5rem">Plano de Ação Sugerido</h2>
    <div class="phase">
        <div class="phase-title">Fase 1 · Curto Prazo (0–6 meses)</div>
        <ul>
            <li>Mapear categorias não captadas e planejar a abertura de vagas</li>
            <li>Realizar busca ativa de alunos com deficiência para registro do AEE (dupla matrícula)</li>
            <li>Iniciar o cumprimento das 5 condicionalidades do VAAR (incluindo a Cond. IV, exclusiva da rede estadual)</li>
            <li>Implementar a BNCC Computação (obrigatória a partir de 2026) em todas as escolas estaduais</li>
        </ul>
    </div>
    <div class="phase">
        <div class="phase-title">Fase 2 · Médio Prazo (6–12 meses)</div>
        <ul>
            <li>Converter {fmt_n(inep.get('em_parcial_est',0))} alunos do Ensino Médio de parcial para integral</li>
            <li>Aderir ao PETI (Programa Escola em Tempo Integral) para as escolas estaduais</li>
            <li>Reclassificar escolas rurais e em terras indígenas ou quilombolas</li>
            <li>Expandir a oferta de Educação Profissional Técnica integrada ao Ensino Médio</li>
        </ul>
    </div>
    <div class="phase">
        <div class="phase-title">Fase 3 · Longo Prazo (12–24 meses)</div>
        <ul>
            <li>Manter monitoramento contínuo do Censo Escolar</li>
            <li>Implantar painel de acompanhamento dos indicadores do VAAR</li>
            <li>Ampliar o tempo integral de {pot['t6']['pct_integral']:.1f}% para a meta de 50% do PNE</li>
        </ul>
    </div>

    <div class="page-footer">Instituto i10 — Orquestrando o Futuro da Educação Pública &nbsp;&nbsp; 6 / 7</div>
</div>

<!-- PAGE 7: CTA -->
<div class="page">
    <div class="cta">
        <h3>O Instituto i10 pode implementar este plano para a Rede Estadual de {nome_estado}</h3>
        <p>Tecnologia · Dados · Pessoas · Suporte Jurídico para maximizar sua captação FUNDEB</p>
        <div class="services">
            <div class="svc">Plataforma de Inteligência</div>
            <div class="svc">Compliance VAAR</div>
            <div class="svc">Busca Ativa AEE</div>
            <div class="svc">Suporte Jurídico</div>
            <div class="svc">BNCC Computação</div>
        </div>
        <div class="email">contato@institutoi10.org.br</div>
    </div>

    <div class="page-footer">Instituto i10 — Orquestrando o Futuro da Educação Pública &nbsp;&nbsp; 7 / 7</div>
</div>

</body></html>"""

    return html


def load_state_comparison(uf):
    df = pd.read_excel(XLSX_RECEITA, header=8, sheet_name=0)
    df.columns = ['uf', 'codigo_ibge', 'entidade', 'receita_contribuicao',
                   'compl_vaaf', 'compl_vaat', 'compl_vaar', 'compl_uniao_total', 'total_receitas']
    states = df[df['entidade'].astype(str).str.contains('GOVERNO', case=False, na=False)]

    region_ufs = {
        'PR': ['PR', 'SC', 'RS'],
        'SC': ['PR', 'SC', 'RS'],
        'RS': ['PR', 'SC', 'RS'],
        'SP': ['SP', 'MG', 'RJ', 'ES'],
        'MG': ['SP', 'MG', 'RJ', 'ES'],
        'RJ': ['SP', 'MG', 'RJ', 'ES'],
    }
    compare_ufs = region_ufs.get(uf, ['PR', 'SC', 'RS', 'SP', 'MG'])

    results = []
    for _, row in states.iterrows():
        row_uf = row['uf']
        if row_uf in compare_ufs:
            results.append({
                'uf': row_uf,
                'mat': 0,
                'receita': float(row['total_receitas']),
            })

    wb = openpyxl.load_workbook(str(SINOPSE), read_only=True, data_only=True)
    ws = wb['1.2']
    for r in results:
        state_name = SINOPSE_STATES.get(r['uf'], '')
        row = find_state_row(ws, state_name)
        if row:
            r['mat'] = safe_int(row[7]) + safe_int(row[12])
    wb.close()

    results.sort(key=lambda x: -x['receita'])
    return results


def generate_report(uf):
    uf = uf.upper()
    if uf not in UF_NAMES:
        print(f"ERRO: UF '{uf}' não reconhecida")
        return

    nome = UF_NAMES[uf]
    print(f"\n{'='*60}")
    print(f"  Gerando relatório: REDE ESTADUAL — {nome}")
    print(f"{'='*60}")

    print("  [1/5] Carregando receita FUNDEB...")
    receita = load_receita(uf)
    if not receita:
        return

    print("  [2/5] Carregando dados VAAR...")
    vaar = load_vaar(uf)

    print("  [3/5] Extraindo matrículas INEP Sinopse...")
    inep = load_inep(uf)

    print("  [4/5] Calculando potencial T1-T6...")
    cats, total_mat = build_cats_state(inep)
    pot = calc_potencial(cats, receita, vaar, inep)

    print(f"\n  --- RESULTADOS {uf} ---")
    print(f"  Receita FUNDEB:     {fmt(pot['receita_fundeb'])}")
    print(f"  Potencial Total:    {fmt(pot['potencial_total'])}")
    print(f"  Ganho %:            +{pot['pct_ganho']:.1f}%")
    print(f"  Matrículas Estadual: {fmt_n(pot['total_mat_est'])}")
    print(f"  Categorias ativas:  {pot['cats_ativas']}/{pot['cats_total']}")
    print(f"  EM Integral:        {pot['t6']['pct_integral']:.1f}%")
    print(f"  Recebe VAAR:        {'SIM' if pot['recebe_vaar'] else 'NÃO'}")
    print(f"  T1 (VAAF):          {fmt(pot['t1']['total'])}")
    print(f"  T2 (Integral):      {fmt(pot['t2']['total'])}")
    print(f"  T3 (AEE):           {fmt(pot['t3']['total'])}")
    print(f"  T4 (Localidade):    {fmt(pot['t4']['total'])}")
    print(f"  T5 (VAAR):          {fmt(pot['t5']['total'])}")
    print(f"  T6 (EC 135):        {fmt(pot['t6']['total'])}")

    print("  [5/5] Gerando PDF...")
    html = build_html(uf, receita, vaar, inep, cats, pot)

    out_pdf = OUT_DIR / f"REDE_ESTADUAL_{uf}.pdf"
    out_html = OUT_DIR / f"REDE_ESTADUAL_{uf}.html"

    with open(out_html, 'w', encoding='utf-8') as f:
        f.write(html)

    HTML(string=html).write_pdf(str(out_pdf))

    print(f"\n  PDF: {out_pdf}")
    print(f"  HTML: {out_html}")
    return pot


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 gerar_relatorio_estadual.py SC PR [SP ...]")
        sys.exit(1)

    ufs = sys.argv[1:]
    for uf in ufs:
        generate_report(uf)

    print(f"\nRelatórios salvos em: {OUT_DIR}")
